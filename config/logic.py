import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import logging
import traceback

def process_csv(file_path, start_date=None, end_date=None):
    try:
        logging.info("Iniciando o processamento do arquivo CSV...")
        print("Iniciando o processamento do arquivo CSV...")
    
        # Ler o arquivo CSV completo
        df = pd.read_csv(
            file_path,
            encoding='ISO-8859-1',
            delimiter=';',
            on_bad_lines='skip'
        )
        # Verificar qual nome da coluna de data está presente e padronizar
        if 'Data_de_Impressão' in df.columns:
            df = df.rename(columns={'Data_de_Impressão': 'Data_de_Impressao'})
        elif 'Data_de_Impressao' not in df.columns:
            raise ValueError("Coluna de data não encontrada!")

        # Filtrar as colunas necessárias
        df = df[['Nome_Completo', 'Paginas_Color', 'Paginas_Mono', 'Data_de_Impressao']]

        # Remover espaços em branco dos nomes e da coluna de data
        df['Nome_Completo'] = df['Nome_Completo'].str.strip().str.lower()
        df['Data_de_Impressao'] = df['Data_de_Impressao'].str.strip()
        
        # Converter a coluna 'Data_de_Impressão' para o tipo datetime
        df['Data_de_Impressao'] = pd.to_datetime(df['Data_de_Impressao'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
        
        # Ajustar o start_date e end_date para incluir o início e o fim do dia
        if start_date is not None:
            start_date = pd.to_datetime(start_date).replace(hour=0, minute=0, second=0)  # 00:00:00 do dia
        if end_date is not None:
            end_date = pd.to_datetime(end_date).replace(hour=23, minute=59, second=59)  # 23:59:59 do dia
        
        # Filtrar os dados com base nas datas fornecidas
        if start_date is not None:
            df = df[df['Data_de_Impressao'] >= start_date]
        if end_date is not None:
            df = df[df['Data_de_Impressao'] <= end_date]
        
        # Garantir que as colunas numéricas estejam no formato correto e preencher valores ausentes com 0
        df['Paginas_Color'] = pd.to_numeric(df['Paginas_Color'], errors='coerce').fillna(0)
        df['Paginas_Mono'] = pd.to_numeric(df['Paginas_Mono'], errors='coerce').fillna(0)
        
        # Criar a coluna 'pagina' como a soma de 'Paginas_Mono' e 'Paginas_Color'
        df['pagina'] = df['Paginas_Color'] + df['Paginas_Mono']
        
        # Renomear as colunas
        df = df.rename(columns={
            'Nome_Completo': 'Nome',
            'Paginas_Color': 'Colorido',
            'Paginas_Mono': 'P&B',
            'pagina': 'Pagina'
        })
        
        # Agrupar pela coluna 'Nome' e somar as colunas numéricas
        df_agrupado = df.groupby('Nome', as_index=False).agg({
            'Colorido': 'sum',
            'P&B': 'sum',
            'Pagina': 'sum'
        })
        
        # Adicionar a linha de total
        total_row = pd.DataFrame(df_agrupado[['Colorido', 'P&B', 'Pagina']].sum()).T
        total_row['Nome'] = 'Total'
        df_agrupado = pd.concat([df_agrupado, total_row], ignore_index=True)
        
        return df_agrupado
    except Exception: 
        error = traceback.format_exc()
        logging.error(error)
        print(error)
    finally:
        logging.info("Fim do processamento do arquivo CSV.")
        print("Fim do processamento do arquivo CSV.")

def save_to_excel(df_agrupado, output_file):
    try:
        logging.info("Salvando o resultado em um arquivo Excel...")
        print("Salvando o resultado em um arquivo Excel...")
    
        # Salvando o resultado em um arquivo Excel
        df_agrupado.to_excel(output_file, index=False)

        # Aplicar bordas e formatação às células no Excel
        wb = load_workbook(output_file)
        ws = wb.active

        # Definir bordas
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Definir fonte em negrito
        bold_font = Font(bold=True)

        # Aplicar bordas e fonte em negrito ao cabeçalho
        for cell in ws[1]:
            cell.border = border
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar bordas e fonte em negrito à linha de total
        total_row_index = len(df_agrupado) + 1
        for cell in ws[total_row_index]:
            cell.border = border
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar bordas a todas as outras células da planilha
        for row in ws.iter_rows(min_row=2, max_row=total_row_index-1, min_col=1, max_col=len(df_agrupado.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Salvar o arquivo com bordas e formatação
        wb.save(output_file)
    except Exception:
        error = traceback.format_exc()
        logging.error(error)
        print(error)
    finally:
        logging.info("Fim da criação do arquivo Excel.")
        print("Fim da criação do arquivo Excel.")